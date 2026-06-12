#!/usr/bin/env python3
"""Build all Olé Quiniela data files from the members spreadsheet.

Outputs (data/):
  teams.json        48-team table (code, name, flag, group, rank, conf)
  matches.json      72 group-stage fixtures (stable ids, group, date, venue)
  predictions.json  every member's pick per fixture + 2 bonus picks
  results.json      actual results (created once as a blank template; an
                    existing file is preserved so live scores aren't wiped)

Usage:
  python3 scripts/build_data.py [members.xlsx] [sibling_schedule.json]

The sibling schedule (wc-2026-draft/data/matches.json) is only read to copy
the official kickoff order / dates / venues; if absent we fall back to the
spreadsheet's own match order.
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
DATA = ROOT / "data"

# ---------------------------------------------------------------------------
# Canonical 48-team table. `aliases` cover every spelling we must resolve:
# quiniela short names + typos, the sibling project's names, and the
# football-data.org feed names. Display `name` is what the site shows.
# ---------------------------------------------------------------------------
TEAMS = [
    {"code": "ESP", "name": "Spain",         "flag": "🇪🇸", "rank": 1,  "conf": "UEFA",     "group": "H"},
    {"code": "FRA", "name": "France",        "flag": "🇫🇷", "rank": 2,  "conf": "UEFA",     "group": "I"},
    {"code": "ARG", "name": "Argentina",     "flag": "🇦🇷", "rank": 3,  "conf": "CONMEBOL", "group": "J"},
    {"code": "ENG", "name": "England",       "flag": "🏴\U000e0067\U000e0062\U000e0065\U000e006e\U000e0067\U000e007f", "rank": 4, "conf": "UEFA", "group": "L"},
    {"code": "BRA", "name": "Brazil",        "flag": "🇧🇷", "rank": 5,  "conf": "CONMEBOL", "group": "C"},
    {"code": "POR", "name": "Portugal",      "flag": "🇵🇹", "rank": 6,  "conf": "UEFA",     "group": "K"},
    {"code": "GER", "name": "Germany",       "flag": "🇩🇪", "rank": 7,  "conf": "UEFA",     "group": "E"},
    {"code": "NED", "name": "Netherlands",   "flag": "🇳🇱", "rank": 8,  "conf": "UEFA",     "group": "F"},
    {"code": "MAR", "name": "Morocco",       "flag": "🇲🇦", "rank": 9,  "conf": "CAF",      "group": "C"},
    {"code": "NOR", "name": "Norway",        "flag": "🇳🇴", "rank": 10, "conf": "UEFA",     "group": "I"},
    {"code": "BEL", "name": "Belgium",       "flag": "🇧🇪", "rank": 11, "conf": "UEFA",     "group": "G"},
    {"code": "COL", "name": "Colombia",      "flag": "🇨🇴", "rank": 12, "conf": "CONMEBOL", "group": "K"},
    {"code": "SEN", "name": "Senegal",       "flag": "🇸🇳", "rank": 13, "conf": "CAF",      "group": "I"},
    {"code": "CRO", "name": "Croatia",       "flag": "🇭🇷", "rank": 14, "conf": "UEFA",     "group": "L"},
    {"code": "JPN", "name": "Japan",         "flag": "🇯🇵", "rank": 15, "conf": "AFC",      "group": "F", "aliases": ["Japana"]},
    {"code": "MEX", "name": "Mexico",        "flag": "🇲🇽", "rank": 16, "conf": "CONCACAF", "group": "A"},
    {"code": "USA", "name": "USA",           "flag": "🇺🇸", "rank": 17, "conf": "CONCACAF", "group": "D", "aliases": ["United States", "United States of America"]},
    {"code": "URU", "name": "Uruguay",       "flag": "🇺🇾", "rank": 18, "conf": "CONMEBOL", "group": "H"},
    {"code": "SUI", "name": "Switzerland",   "flag": "🇨🇭", "rank": 19, "conf": "UEFA",     "group": "B"},
    {"code": "IRN", "name": "Iran",          "flag": "🇮🇷", "rank": 20, "conf": "AFC",      "group": "G", "aliases": ["IR Iran", "Iran IR"]},
    {"code": "TUR", "name": "Türkiye",       "flag": "🇹🇷", "rank": 21, "conf": "UEFA",     "group": "D", "aliases": ["Turkey", "Turkiye", "Turkoye"]},
    {"code": "ECU", "name": "Ecuador",       "flag": "🇪🇨", "rank": 22, "conf": "CONMEBOL", "group": "E"},
    {"code": "AUT", "name": "Austria",       "flag": "🇦🇹", "rank": 23, "conf": "UEFA",     "group": "J"},
    {"code": "KOR", "name": "South Korea",   "flag": "🇰🇷", "rank": 24, "conf": "AFC",      "group": "A", "aliases": ["Korea", "Korea Republic", "Korea, South", "Republic of Korea"]},
    {"code": "AUS", "name": "Australia",     "flag": "🇦🇺", "rank": 25, "conf": "AFC",      "group": "D"},
    {"code": "ALG", "name": "Algeria",       "flag": "🇩🇿", "rank": 26, "conf": "CAF",      "group": "J"},
    {"code": "EGY", "name": "Egypt",         "flag": "🇪🇬", "rank": 27, "conf": "CAF",      "group": "G"},
    {"code": "CAN", "name": "Canada",        "flag": "🇨🇦", "rank": 28, "conf": "CONCACAF", "group": "B"},
    {"code": "PAN", "name": "Panama",        "flag": "🇵🇦", "rank": 29, "conf": "CONCACAF", "group": "L"},
    {"code": "CIV", "name": "Côte d'Ivoire", "flag": "🇨🇮", "rank": 30, "conf": "CAF",      "group": "E", "aliases": ["Ivory Coast", "Cote d'Ivoire", "Cote d'Ivore", "Cote dIvoire", "Cote dIvore"]},
    {"code": "SWE", "name": "Sweden",        "flag": "🇸🇪", "rank": 31, "conf": "UEFA",     "group": "F"},
    {"code": "PAR", "name": "Paraguay",      "flag": "🇵🇾", "rank": 32, "conf": "CONMEBOL", "group": "D"},
    {"code": "CZE", "name": "Czechia",       "flag": "🇨🇿", "rank": 33, "conf": "UEFA",     "group": "A", "aliases": ["Czech Republic"]},
    {"code": "SCO", "name": "Scotland",      "flag": "🏴\U000e0067\U000e0062\U000e0073\U000e0063\U000e0074\U000e007f", "rank": 34, "conf": "UEFA", "group": "C"},
    {"code": "TUN", "name": "Tunisia",       "flag": "🇹🇳", "rank": 35, "conf": "CAF",      "group": "F"},
    {"code": "COD", "name": "DR Congo",      "flag": "🇨🇩", "rank": 36, "conf": "CAF",      "group": "K", "aliases": ["Congo", "Congo DR", "Democratic Republic of the Congo", "Congo (DR)"]},
    {"code": "UZB", "name": "Uzbekistan",    "flag": "🇺🇿", "rank": 37, "conf": "AFC",      "group": "K"},
    {"code": "QAT", "name": "Qatar",         "flag": "🇶🇦", "rank": 38, "conf": "AFC",      "group": "B"},
    {"code": "IRQ", "name": "Iraq",          "flag": "🇮🇶", "rank": 39, "conf": "AFC",      "group": "I"},
    {"code": "RSA", "name": "South Africa",  "flag": "🇿🇦", "rank": 40, "conf": "CAF",      "group": "A"},
    {"code": "KSA", "name": "Saudi Arabia",  "flag": "🇸🇦", "rank": 41, "conf": "AFC",      "group": "H"},
    {"code": "JOR", "name": "Jordan",        "flag": "🇯🇴", "rank": 42, "conf": "AFC",      "group": "J"},
    {"code": "BIH", "name": "Bosnia & Herzegovina", "flag": "🇧🇦", "rank": 43, "conf": "UEFA", "group": "B", "aliases": ["Bosnia", "Bosnia and Herzegovina", "Bosnia-Herzegovina", "Bosnia and Herz."]},
    {"code": "CPV", "name": "Cabo Verde",    "flag": "🇨🇻", "rank": 44, "conf": "CAF",      "group": "H", "aliases": ["Cape Verde", "Cape Verde Islands"]},
    {"code": "GHA", "name": "Ghana",         "flag": "🇬🇭", "rank": 45, "conf": "CAF",      "group": "L"},
    {"code": "CUW", "name": "Curaçao",       "flag": "🇨🇼", "rank": 46, "conf": "CONCACAF", "group": "E", "aliases": ["Curacao"]},
    {"code": "HAI", "name": "Haiti",         "flag": "🇭🇹", "rank": 47, "conf": "CONCACAF", "group": "C"},
    {"code": "NZL", "name": "New Zealand",   "flag": "🇳🇿", "rank": 48, "conf": "OFC",      "group": "G"},
]

TIE_TOKENS = {"tie", "draw", "empate", "t", "x"}


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def nk(s):
    """Normalised lookup key."""
    s = strip_accents(s or "").lower().strip()
    s = s.replace("’", "'").replace("`", "'").replace("'", "")
    s = re.sub(r"[^a-z0-9 ]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# alias key -> code
_INDEX = {}
for t in TEAMS:
    _INDEX[nk(t["name"])] = t["code"]
    _INDEX[nk(t["code"])] = t["code"]
    for a in t.get("aliases", []):
        _INDEX[nk(a)] = t["code"]
_BY_CODE = {t["code"]: t for t in TEAMS}


def edist(a, b):
    if a == b:
        return 0
    if not a or not b:
        return len(a or b)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def resolve_code(token, allowed=None, warnings=None):
    """Resolve a raw team token to a team code. `allowed` restricts to a
    match's two participants (with typo correction)."""
    key = nk(token)
    if not key:
        return None
    code = _INDEX.get(key)
    if code is None:
        # fuzzy against whole table
        best, bd = None, 99
        for k, c in _INDEX.items():
            d = edist(key, k)
            if d < bd:
                best, bd, code_best = c, d, c
        code = best if bd <= 2 else None
    if allowed is None:
        if code is None and warnings is not None:
            warnings.append(f"  UNRESOLVED team token: {token!r}")
        return code
    if code in allowed:
        return code
    # constrain to the match's two teams via name edit distance
    best, bd = None, 99
    for c in allowed:
        d = edist(key, nk(_BY_CODE[c]["name"]))
        for a in _BY_CODE[c].get("aliases", []):
            d = min(d, edist(key, nk(a)))
        if d < bd:
            best, bd = c, d
    if bd <= 3:
        if code != best and warnings is not None:
            warnings.append(f"  fixed {token!r} -> {_BY_CODE[best]['name']} ({'/'.join(_BY_CODE[c]['name'] for c in allowed)})")
        return best
    if warnings is not None:
        warnings.append(f"  UNRESOLVED {token!r} in match {'/'.join(_BY_CODE[c]['name'] for c in allowed)}")
    return code


SCORE_RE = re.compile(r"(\d+)\s*[-–—:]\s*(\d+)")


def parse_pick(cell, home_code, away_code, warnings):
    if cell is None:
        return None
    raw = str(cell).strip()
    if not raw:
        return None
    if ":" in raw and not SCORE_RE.match(raw):
        wtok, stok = raw.split(":", 1)
    else:
        m = SCORE_RE.search(raw)
        if m and m.start() > 0:
            wtok, stok = raw[:m.start()], raw[m.start():]
        elif ":" in raw:
            wtok, stok = raw.split(":", 1)
        else:
            wtok, stok = raw, ""
    if nk(wtok) in TIE_TOKENS:
        code = "TIE"
    else:
        code = resolve_code(wtok, allowed={home_code, away_code}, warnings=warnings)
    m = SCORE_RE.search(stok or "")
    a = int(m.group(1)) if m else None
    b = int(m.group(2)) if m else None

    if code == "TIE":
        result, hg, ag, team = "T", a, b, ""
    elif code == home_code:
        result, hg, ag, team = "H", a, b, home_code
    elif code == away_code:
        result, hg, ag, team = "A", (b if a is not None else None), (a if a is not None else None), away_code
    else:
        result, hg, ag, team = "?", a, b, (code or "")
    return {"result": result, "hg": hg, "ag": ag, "team": team, "raw": raw}


def load_schedule(sibling_path):
    """Return the official kickoff-ordered fixture list. Prefers the already
    built data/matches.json (self-contained); otherwise reads the sibling
    wc-2026-draft schedule to bootstrap it the first time."""
    built = DATA / "matches.json"
    if built.exists():
        order = []
        for m in json.loads(built.read_text()):
            order.append({
                "id": m["id"], "group": m.get("group"), "matchday": m.get("matchday"),
                "utcDate": m.get("utcDate"), "venue": m.get("venue"),
                "homeCode": m["homeCode"], "awayCode": m["awayCode"],
            })
        return order
    order = []
    if sibling_path and Path(sibling_path).exists():
        sib = json.loads(Path(sibling_path).read_text())
        for m in sib["matches"]:
            order.append({
                "id": m["id"], "group": m.get("group"), "matchday": m.get("matchday"),
                "utcDate": m.get("utcDate"), "venue": m.get("venue"),
                "homeCode": resolve_code(m["home"]), "awayCode": resolve_code(m["away"]),
            })
    return order


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/Users/felipeortiz/Downloads/2026_Quiniela_Members.xlsx")
    sibling = sys.argv[2] if len(sys.argv) > 2 else "/Users/felipeortiz/Workspace/wc-2026-draft/data/matches.json"
    warnings = []

    schedule = load_schedule(sibling)
    by_pair = {(s["homeCode"], s["awayCode"]): s for s in schedule}
    sched_by_id = {s["id"]: s for s in schedule}

    wb = openpyxl.load_workbook(src, data_only=True)
    members, picks, specials = [], {}, {}
    seen_match_ids = []

    for si, ws in enumerate(wb.worksheets):
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        member_cols = [(c, str(v).strip()) for c, v in enumerate(header, 1)
                       if c >= 2 and v and str(v).strip()]
        for _, name in member_cols:
            if name not in members:
                members.append(name); picks[name] = {}; specials[name] = {}

        for r in range(2, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not label or not str(label).strip():
                continue
            label = str(label).strip(); low = label.lower()
            if "top goal scorer" in low:
                for c, name in member_cols:
                    v = ws.cell(r, c).value
                    specials[name]["topScorer"] = (str(v).strip() if v else "")
                continue
            if "least goal" in low:
                for c, name in member_cols:
                    v = ws.cell(r, c).value
                    specials[name]["leastConceded"] = resolve_code(str(v), warnings=warnings) if v else ""
                continue
            if " vs" not in low:
                continue

            parts = re.split(r"\s+vs\.?\s+", label, flags=re.I)
            if len(parts) != 2:
                warnings.append(f"  bad label {label!r}"); continue
            hc = resolve_code(parts[0], warnings=warnings)
            ac = resolve_code(parts[1], warnings=warnings)
            sched = by_pair.get((hc, ac)) or by_pair.get((ac, hc))
            if not sched:
                warnings.append(f"  no schedule match for {label!r} ({hc} v {ac})"); continue
            mid = sched["id"]
            # use the OFFICIAL home/away orientation for score storage
            ohc, oac = sched["homeCode"], sched["awayCode"]
            if mid not in seen_match_ids:
                seen_match_ids.append(mid)
            for c, name in member_cols:
                pk = parse_pick(ws.cell(r, c).value, ohc, oac, warnings)
                if pk:
                    picks[name][mid] = pk

    # ---- matches.json in official kickoff order ----
    matches = []
    for s in schedule:
        matches.append({
            "id": s["id"], "group": s["group"], "matchday": s["matchday"],
            "utcDate": s["utcDate"], "venue": s["venue"],
            "homeCode": s["homeCode"], "awayCode": s["awayCode"],
        })

    DATA.mkdir(exist_ok=True)
    (DATA / "teams.json").write_text(json.dumps(
        [{k: t[k] for k in ("code", "name", "flag", "group", "rank", "conf")} for t in TEAMS],
        indent=2, ensure_ascii=False) + "\n")
    (DATA / "matches.json").write_text(json.dumps(matches, indent=2, ensure_ascii=False) + "\n")
    (DATA / "predictions.json").write_text(json.dumps(
        {"members": members, "picks": picks, "specials": specials},
        indent=2, ensure_ascii=False) + "\n")

    results_path = DATA / "results.json"
    if not results_path.exists():
        results_path.write_text(json.dumps({
            "updated": None, "topScorer": None, "leastConceded": None,
            "overrides": [],
            "matches": {m["id"]: None for m in matches},
        }, indent=2, ensure_ascii=False) + "\n")

    # ---- report ----
    print(f"members : {len(members)}")
    print(f"matches mapped : {len(seen_match_ids)} / {len(matches)} scheduled")
    pc = [len(picks[m]) for m in members]
    print(f"picks/member   : min {min(pc)} max {max(pc)}")
    # validate every member has every scheduled match
    missing = {m: [mid for mid in sched_by_id if mid not in picks[m]] for m in members}
    missing = {m: v for m, v in missing.items() if v}
    if missing:
        print("MISSING PICKS:")
        for m, v in list(missing.items())[:10]:
            print(f"  {m}: {len(v)} missing -> {v[:5]}")
    if warnings:
        print(f"\n=== {len(warnings)} warnings ===")
        seen = set()
        for w in warnings:
            if w not in seen:
                print(w); seen.add(w)
    else:
        print("\nNo warnings.")


if __name__ == "__main__":
    main()
